from http.server import BaseHTTPRequestHandler
import json
import os
from pathlib import Path
from tempfile import NamedTemporaryFile

import openpyxl
from openpyxl.cell.cell import MergedCell


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def detect_layout(ws):
    row2_col1 = (clean_text(ws.cell(2, 1).value) or "").lower()
    row2_col2 = (clean_text(ws.cell(2, 2).value) or "").lower()
    if "item" in row2_col1 and "description" in row2_col2:
        return {
            "item_col": 1,
            "label_col": 2,
            "first_slot_col": 3,
        }
    return {
        "item_col": None,
        "label_col": 1,
        "first_slot_col": 2,
    }


def find_block_starts(ws, first_slot_col):
    starts = []
    for col in range(first_slot_col, ws.max_column + 1):
        if clean_text(ws.cell(1, col).value):
            starts.append(col)
    if not starts:
        for col in range(first_slot_col, ws.max_column + 1):
            header = (clean_text(ws.cell(2, col).value) or "").lower()
            if header == "asin":
                starts.append(col)
    return starts


def field_key(header):
    lower = (header or "").strip().lower()
    if lower == "asin":
        return "asin"
    if lower == "link":
        return "link"
    if "comment" in lower:
        return "comment"
    if "ranking" in lower or lower == "bsr" or "bsr" in lower:
        return "bsr"
    if "bewertungszahl" in lower or "review count" in lower:
        return "reviewCount"
    if lower == "bewertungen" or "reviews rating" in lower or "rating" in lower:
        return "rating"
    if "preis" in lower or lower == "price":
        return "price"
    return None


def parse_sheet_map(ws):
    layout = detect_layout(ws)
    starts = find_block_starts(ws, layout["first_slot_col"])
    blocks = []
    for index, start in enumerate(starts):
        end = starts[index + 1] - 1 if index + 1 < len(starts) else ws.max_column
        headers = {}
        for col in range(start, end + 1):
            key = field_key(clean_text(ws.cell(2, col).value) or "")
            if key and key not in headers:
                headers[key] = col
        if "asin" not in headers:
            continue
        blocks.append({"headers": headers})

    rows = []
    for row_idx in range(3, ws.max_row + 1):
        label = clean_text(ws.cell(row_idx, layout["label_col"]).value)
        item_number = clean_text(ws.cell(row_idx, layout["item_col"]).value) if layout["item_col"] else None
        has_any = False
        for block in blocks:
            asin_col = block["headers"].get("asin")
            if asin_col and clean_text(ws.cell(row_idx, asin_col).value):
                has_any = True
                break
        if label or item_number or has_any:
            rows.append(
                {
                    "row_idx": row_idx,
                    "label": label or "",
                    "itemNumber": item_number,
                    "blocks": blocks,
                }
            )
    return rows


def top_left_of_merged_range(ws, row_idx, col_idx):
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row_idx <= merged_range.max_row
            and merged_range.min_col <= col_idx <= merged_range.max_col
        ):
            return merged_range.min_row, merged_range.min_col
    return row_idx, col_idx


def write_value(ws, row_idx, col_idx, value):
    if not col_idx:
        return
    target_row, target_col = top_left_of_merged_range(ws, row_idx, col_idx)
    cell = ws.cell(row=target_row, column=target_col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def build_workbook(template_path: str, payload: dict) -> bytes:
    wb = openpyxl.load_workbook(template_path)
    by_category = {entry["label"]: entry for entry in payload.get("categories", [])}

    for ws in wb.worksheets:
        category = by_category.get(ws.title)
        if not category:
            continue

        sheet_rows = parse_sheet_map(ws)
        payload_rows = category.get("rows", [])

        for row_map, row_payload in zip(sheet_rows, payload_rows):
            blocks = row_map["blocks"]
            slots = row_payload.get("slots", [])

            for block, slot in zip(blocks, slots):
                headers = block["headers"]
                current = slot.get("current") or {}

                if headers.get("price"):
                    write_value(ws, row_map["row_idx"], headers.get("price"), current.get("price"))
                if headers.get("bsr"):
                    write_value(ws, row_map["row_idx"], headers.get("bsr"), current.get("bsr"))
                if headers.get("reviewCount"):
                    write_value(ws, row_map["row_idx"], headers.get("reviewCount"), current.get("reviewCount"))
                if headers.get("rating"):
                    write_value(ws, row_map["row_idx"], headers.get("rating"), current.get("rating"))

    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb.save(tmp_path)
        with open(tmp_path, "rb") as f:
            return f.read()
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            content_length = int(self.headers.get("content-length", "0"))
            raw = self.rfile.read(content_length)
            body = json.loads(raw.decode("utf-8"))

            payload = body.get("payload")
            template_rel = body.get("templatePath")
            filename_base = body.get("filenameBase") or "competitor_analysis_export"

            if not payload or not template_rel:
                self.send_response(400)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps({
                    "ok": False,
                    "error": "payload oder templatePath fehlt"
                }).encode("utf-8"))
                return

            root_dir = Path(os.getcwd())
            template_path = root_dir / template_rel

            if not template_path.exists():
                self.send_response(400)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps({
                    "ok": False,
                    "error": f"Template nicht gefunden: {template_rel}"
                }).encode("utf-8"))
                return

            workbook_bytes = build_workbook(str(template_path), payload)

            self.send_response(200)
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header(
                "Content-Disposition",
                f'attachment; filename="{filename_base}.xlsx"'
            )
            self.end_headers()
            self.wfile.write(workbook_bytes)

        except Exception as e:
            self.send_response(500)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            self.wfile.write(json.dumps({
                "ok": False,
                "error": str(e),
            }).encode("utf-8"))