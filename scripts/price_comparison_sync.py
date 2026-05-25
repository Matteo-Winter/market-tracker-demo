#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json
import re
import sys
from pathlib import Path

import openpyxl

YELLOW_MARKERS = {"FFFFFF00", "FFFF00", "FFF2CC", "FFD966"}

def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None

def normalize_asin(value):
    text = clean_text(value)
    if not text:
        return None
    text = text.upper()
    return text if re.fullmatch(r"[A-Z0-9]{10}", text) else None

def slugify(value):
    text = (value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or "category"

def parse_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None

def parse_int(value):
    parsed = parse_float(value)
    if parsed is None:
        return None
    return int(round(parsed))

def color_is_yellow(cell):
    fill = getattr(cell, "fill", None)
    if not fill or not fill.fill_type:
        return False

    candidates = []
    for color in [getattr(fill, "fgColor", None), getattr(fill, "start_color", None), getattr(fill, "end_color", None)]:
        if not color:
            continue
        if getattr(color, "type", None) == "rgb" and getattr(color, "rgb", None):
            candidates.append(str(color.rgb).upper())
        if getattr(color, "type", None) == "indexed" and getattr(color, "indexed", None) is not None:
            candidates.append(f"INDEXED:{color.indexed}")

    if any(any(marker in value for marker in YELLOW_MARKERS) for value in candidates):
        return True

    # Fallback: jede bewusst gesetzte Fill-Farbe ungleich Standard als manuell behandeln.
    return fill.fill_type == "solid" and bool(candidates)

def detect_layout(ws):
    row2_col1 = (clean_text(ws.cell(2, 1).value) or "").lower()
    row2_col2 = (clean_text(ws.cell(2, 2).value) or "").lower()
    if "item" in row2_col1 and "description" in row2_col2:
        return {
            "item_col": 1,
            "label_col": 2,
            "first_slot_col": 3,
        }
    return {
        "item_col": None,
        "label_col": 1,
        "first_slot_col": 2,
    }

def find_block_starts(ws, first_slot_col):
    starts = []
    for col in range(first_slot_col, ws.max_column + 1):
        if clean_text(ws.cell(1, col).value):
            starts.append(col)
    if not starts:
        for col in range(first_slot_col, ws.max_column + 1):
            header = (clean_text(ws.cell(2, col).value) or "").lower()
            if header == "asin":
                starts.append(col)
    return starts

def field_key(header):
    lower = (header or "").strip().lower()
    if lower == "asin":
        return "asin"
    if lower == "link":
        return "link"
    if "comment" in lower:
        return "comment"
    if "ranking" in lower or lower == "bsr" or "bsr" in lower:
        return "bsr"
    if "bewertungszahl" in lower or "review count" in lower:
        return "reviewCount"
    if lower == "bewertungen" or "reviews rating" in lower or "rating" in lower:
        return "rating"
    if "preis" in lower or lower == "price":
        return "price"
    if "umsatz" in lower or "asin revenue" in lower:
        return "asinRevenue"
    return None

def parse_sheet(ws):
    layout = detect_layout(ws)
    starts = find_block_starts(ws, layout["first_slot_col"])
    if not starts:
        return None

    blocks = []
    for index, start in enumerate(starts):
        end = starts[index + 1] - 1 if index + 1 < len(starts) else ws.max_column
        headers = {}
        for col in range(start, end + 1):
            key = field_key(clean_text(ws.cell(2, col).value) or "")
            if key and key not in headers:
                headers[key] = col
        if "asin" not in headers:
            continue
        blocks.append(
            {
                "brand": clean_text(ws.cell(1, start).value),
                "start": start,
                "end": end,
                "headers": headers,
            }
        )

    if not blocks:
        return None

    rows = []
    for row_idx in range(3, ws.max_row + 1):
        label = clean_text(ws.cell(row_idx, layout["label_col"]).value)
        item_number = clean_text(ws.cell(row_idx, layout["item_col"]).value) if layout["item_col"] else None

        slots = []
        has_any_slot_content = False

        for block in blocks:
            headers = block["headers"]
            asin = normalize_asin(ws.cell(row_idx, headers["asin"]).value)
            seed = {}

            def read_number(name, integer=False):
                col = headers.get(name)
                if not col:
                    return None
                return parse_int(ws.cell(row_idx, col).value) if integer else parse_float(ws.cell(row_idx, col).value)

            price = read_number("price")
            bsr = read_number("bsr", integer=True)
            review_count = read_number("reviewCount", integer=True)
            rating = read_number("rating")
            asin_revenue = read_number("asinRevenue")
            comment = clean_text(ws.cell(row_idx, headers["comment"]).value) if headers.get("comment") else None
            link = clean_text(ws.cell(row_idx, headers["link"]).value) if headers.get("link") else None

            if price is not None:
                seed["price"] = price
            if bsr is not None:
                seed["bsr"] = bsr
            if review_count is not None:
                seed["reviewCount"] = review_count
            if rating is not None:
                seed["rating"] = rating
            if asin_revenue is not None:
                seed["asinRevenue"] = asin_revenue

            yellow = any(color_is_yellow(ws.cell(row_idx, col)) for col in range(block["start"], block["end"] + 1))
            manual_hint = yellow or (asin is None and any(v is not None for v in [comment, link, price, bsr, review_count, rating, asin_revenue]))

            slot = {
                "brand": block["brand"],
                "asin": asin,
                "link": link,
                "comment": comment,
                "manualHint": manual_hint,
            }
            if seed:
                slot["seed"] = seed

            if asin or comment or link or seed or manual_hint:
                has_any_slot_content = True

            slots.append(slot)

        if label or item_number or has_any_slot_content:
            rows.append(
                {
                    "label": label or "",
                    "itemNumber": item_number,
                    "slots": slots,
                }
            )

    if not rows:
        return None

    return {
        "id": ws.title,
        "label": ws.title,
        "rows": rows,
    }

def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: price_comparison_sync.py <template.xlsx> <template.json>")

    template_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    wb = openpyxl.load_workbook(template_path)
    categories = []

    for ws in wb.worksheets:
        parsed = parse_sheet(ws)
        if parsed:
            categories.append(parsed)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(categories, ensure_ascii=False, indent=2), encoding="utf-8")

if __name__ == "__main__":
    main()
