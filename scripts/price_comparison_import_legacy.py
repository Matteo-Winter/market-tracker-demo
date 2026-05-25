#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json
import re
import sys
from pathlib import Path

import openpyxl


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def normalize_asin(value):
    text = clean_text(value)
    if not text:
        return None
    text = text.upper()
    return text if re.fullmatch(r"[A-Z0-9]{10}", text) else None


def parse_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text or text in {"-", "–"}:
        return None

    text = text.replace("€", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        return float(text)
    except Exception:
        return None


def parse_int(value):
    parsed = parse_float(value)
    if parsed is None:
        return None
    return int(round(parsed))


def detect_layout(ws):
    row2_col1 = (clean_text(ws.cell(2, 1).value) or "").lower()
    row2_col2 = (clean_text(ws.cell(2, 2).value) or "").lower()
    if "item" in row2_col1 and "description" in row2_col2:
        return {
            "item_col": 1,
            "label_col": 2,
            "first_slot_col": 3,
        }
    return {
        "item_col": None,
        "label_col": 1,
        "first_slot_col": 2,
    }


def find_block_starts(ws, first_slot_col):
    starts = []
    for col in range(first_slot_col, ws.max_column + 1):
        if clean_text(ws.cell(1, col).value):
            starts.append(col)
    if not starts:
        for col in range(first_slot_col, ws.max_column + 1):
            header = (clean_text(ws.cell(2, col).value) or "").lower()
            if header == "asin":
                starts.append(col)
    return starts


def field_key(header):
    lower = (header or "").strip().lower()
    if lower == "asin":
        return "asin"
    if lower == "link":
        return "link"
    if "comment" in lower:
        return "comment"
    if "ranking" in lower or lower == "bsr" or "bsr" in lower:
        return "bsr"
    if "bewertungszahl" in lower or "review count" in lower:
        return "reviewCount"
    if lower == "bewertungen" or "reviews rating" in lower or "rating" in lower:
        return "rating"
    if "preis" in lower or lower == "price":
        return "price"
    if "umsatz" in lower or "asin revenue" in lower:
        return "asinRevenue"
    return None


def normalize_review_fields(raw_review_count, raw_rating):
    a = parse_float(raw_review_count)
    b = parse_float(raw_rating)

    if a is not None and b is not None:
        if a <= 5 and b > 5:
            return parse_int(b), a
        return parse_int(a) if a > 5 else None, b if b <= 5 else None

    if a is not None:
        if a <= 5:
            return None, a
        return parse_int(a), None

    if b is not None:
        if b <= 5:
            return None, b
        return parse_int(b), None

    return None, None


def parse_sheet(ws):
    layout = detect_layout(ws)
    starts = find_block_starts(ws, layout["first_slot_col"])
    if not starts:
        return []

    blocks = []
    for index, start in enumerate(starts):
        end = starts[index + 1] - 1 if index + 1 < len(starts) else ws.max_column
        headers = {}
        for col in range(start, end + 1):
            key = field_key(clean_text(ws.cell(2, col).value) or "")
            if key and key not in headers:
                headers[key] = col
        if "asin" not in headers:
            continue
        blocks.append(
            {
                "brand": clean_text(ws.cell(1, start).value),
                "headers": headers,
            }
        )

    if not blocks:
        return []

    items = []
    for row_idx in range(3, ws.max_row + 1):
        title = clean_text(ws.cell(row_idx, layout["label_col"]).value)

        for block in blocks:
            headers = block["headers"]
            asin = normalize_asin(ws.cell(row_idx, headers["asin"]).value)
            if not asin:
                continue

            raw_review_count = ws.cell(row_idx, headers["reviewCount"]).value if headers.get("reviewCount") else None
            raw_rating = ws.cell(row_idx, headers["rating"]).value if headers.get("rating") else None
            review_count, rating = normalize_review_fields(raw_review_count, raw_rating)

            items.append(
                {
                    "asin": asin,
                    "title": title,
                    "brand": block["brand"],
                    "category": ws.title,
                    "price": parse_float(ws.cell(row_idx, headers["price"]).value) if headers.get("price") else None,
                    "bsr": parse_int(ws.cell(row_idx, headers["bsr"]).value) if headers.get("bsr") else None,
                    "reviewCount": review_count,
                    "rating": rating,
                    "asinRevenue": parse_float(ws.cell(row_idx, headers["asinRevenue"]).value) if headers.get("asinRevenue") else None,
                }
            )

    return items


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: price_comparison_import_legacy.py <historic.xlsx>")

    workbook_path = Path(sys.argv[1])
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    items = []

    for ws in wb.worksheets:
        items.extend(parse_sheet(ws))

    sys.stdout.write(json.dumps({"items": items}, ensure_ascii=False))


if __name__ == "__main__":
    main()
